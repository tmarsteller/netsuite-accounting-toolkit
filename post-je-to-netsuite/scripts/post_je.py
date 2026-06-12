"""
Translate a handoff-format JSON into the exact `ns_createRecord` payload.

This module does NOT call the MCP tool directly — it returns the dict that
the Claude Code agent passes to:

    mcp__<your-netsuite-connector-id>__ns_createRecord

That keeps the OAuth/authentication concern out of the Python layer (MCP
handles it) and matches the <your-automation> pattern.

Public API:
    to_ns_payload(handoff: dict) -> dict
    deep_link(record_type: str, internal_id: str, account_id: str = "...") -> str
    from_excel(path: str) -> dict          # convenience: read Summary-for-JE tab
    from_legacy_payload(legacy: dict) -> dict  # convert <your-automation>-style output
"""

from __future__ import annotations

import os
from decimal import Decimal
from typing import Any

SCHEMA_VERSION = "v1"
DEFAULT_CUSTOM_FORM = None   # Set to your org's standard JE form ID, or leave None for NS default
DEFAULT_CURRENCY = "1"

# Hard-coded — these are the Pending Approval guard. Do not parameterize.
_PENDING_APPROVAL = {
    "approved": False,
    "approvalStatus": {"id": "1"},
}


def _ref(value: Any) -> dict | None:
    """Wrap a scalar id into NetSuite's {id: "..."} reference shape."""
    if value is None or value == "":
        return None
    return {"id": str(value).strip()}


def _money(v: Any) -> float | None:
    if v is None or v == "":
        return None
    d = Decimal(str(v)).quantize(Decimal("0.01"))
    return float(d)


def _build_line(ln: dict, is_aicje: bool, header_sub: str) -> dict:
    """Translate one handoff line into a NetSuite line.items[] entry."""
    out: dict[str, Any] = {}
    out["account"] = _ref(ln.get("account"))

    d = _money(ln.get("debit"))
    c = _money(ln.get("credit"))
    if d is not None:
        out["debit"] = d
    if c is not None:
        out["credit"] = c

    if ln.get("memo"):
        out["memo"] = str(ln["memo"])

    for opt_field, key in [
        ("department", "department"),
        ("class", "class"),
        ("location", "location"),
        ("entity", "entity"),
    ]:
        ref = _ref(ln.get(opt_field))
        if ref:
            out[key] = ref

    # Line-level subsidiary handling:
    #   - Normal JE: use lineSubsidiary + dueToFromSubsidiary if the line is
    #     on a different sub than the header (per-line IC pattern)
    #   - AICJE: do NOT set lineSubsidiary on lines; subsidiary allocation is
    #     header-level via subsidiary + toSubsidiaries
    line_sub = ln.get("subsidiary")
    if line_sub and not is_aicje and str(line_sub) != str(header_sub):
        out["lineSubsidiary"] = _ref(line_sub)
        out["dueToFromSubsidiary"] = _ref(line_sub)

    return out


def to_ns_payload(handoff: dict) -> dict:
    """
    Convert a handoff JSON into the {recordType, data} dict for ns_createRecord.

    Hard-codes approved=False and approvalStatus={id:"1"}, regardless of what
    the inbound handoff says.
    """
    # Detect AICJE: presence of to_subsidiaries OR multiple distinct subs across lines
    header_sub = str(handoff.get("subsidiary", "")).strip()
    to_subs_raw = handoff.get("to_subsidiaries") or []
    if isinstance(to_subs_raw, (str, int)):
        to_subs_raw = [to_subs_raw]
    to_subs = [str(s).strip() for s in to_subs_raw if str(s).strip()]

    line_subs = {
        str(ln.get("subsidiary")).strip()
        for ln in handoff.get("lines", [])
        if ln.get("subsidiary")
    }
    distinct_subs = ({header_sub} if header_sub else set()) | set(to_subs) | line_subs

    is_aicje = bool(to_subs) or len(distinct_subs) >= 2

    record_type = "advintercompanyjournalentry" if is_aicje else "journalentry"

    # --- Header ---
    data: dict[str, Any] = {
        "subsidiary": _ref(header_sub),
        "tranDate": handoff["tran_date"],
        "memo": handoff["memo"],
        "customForm": _ref(handoff.get("custom_form", DEFAULT_CUSTOM_FORM)),
        "currency": _ref(handoff.get("currency", DEFAULT_CURRENCY)),
    }

    if is_aicje:
        # NetSuite expects toSubsidiaries as a single ref with comma-separated id
        # for multi-target. Single target: {id: "28"}. Multi: {id: "28,30"}.
        infer_to = sorted(s for s in distinct_subs if s and s != header_sub)
        chosen_to = to_subs if to_subs else infer_to
        if not chosen_to:
            raise ValueError(
                "AICJE classified but no `to_subsidiaries` could be determined."
            )
        data["toSubsidiaries"] = {"id": ",".join(chosen_to)}
        if handoff.get("exchange_rate") is not None:
            data["exchangeRate"] = float(handoff["exchange_rate"])

    # Optional header dims
    for opt, key in [("department", "department"), ("class", "class"), ("location", "location")]:
        ref = _ref(handoff.get(opt))
        if ref:
            data[key] = ref

    # --- Pending Approval guard (hard-coded) ---
    data.update(_PENDING_APPROVAL)

    # --- Lines ---
    data["line"] = {
        "items": [_build_line(ln, is_aicje, header_sub) for ln in handoff["lines"]]
    }

    return {"recordType": record_type, "data": data}


def deep_link(record_type: str, internal_id: str, account_subdomain: str = "") -> str:
    """
    Build a clickable URL to the created record.

    `account_subdomain` is the NetSuite account-specific subdomain (e.g.
    "1234567"); pass it via the NS_ACCOUNT env var if not supplied.
    """
    subdomain = account_subdomain or os.environ.get("NS_ACCOUNT", "<account>")
    path = (
        "advintercompanyjournal.nl"
        if record_type == "advintercompanyjournalentry"
        else "journal.nl"
    )
    return f"https://{subdomain}.app.netsuite.com/app/accounting/transactions/{path}?id={internal_id}"


# --- Convenience: read a "Summary for JE" tab from an Excel workbook ---

def from_excel(path: str, sheet_name: str = "Summary for JE") -> dict:
    """
    Read a JE from an Excel workbook in the <your-automation> / <your-automation> style.

    Expected layout:
      Row 1: column headers (Account | Subsidiary | Debit | Credit | Memo)
      Rows 2..N: line data
      Anywhere: cells labeled 'Memo:' and 'Date:' for the header values.

    This is a best-effort reader — review the result before posting.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl required: `pip install openpyxl`"
        ) from e

    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    header_memo = None
    header_date = None
    header_sub = None
    rows = list(ws.iter_rows(values_only=True))

    # Find header values
    for row in rows[:10]:
        for i, cell in enumerate(row):
            if isinstance(cell, str):
                low = cell.strip().lower()
                if low.startswith("memo") and i + 1 < len(row):
                    header_memo = row[i + 1]
                elif low.startswith("date") and i + 1 < len(row):
                    header_date = row[i + 1]
                elif low.startswith("subsidiary") and i + 1 < len(row):
                    header_sub = row[i + 1]

    # Find the table header row
    table_header_idx = None
    for i, row in enumerate(rows):
        normalized = [str(c).strip().lower() if c else "" for c in row]
        if "account" in normalized and ("debit" in normalized or "credit" in normalized):
            table_header_idx = i
            break

    if table_header_idx is None:
        raise ValueError(
            f"Could not find a line-item header row (Account / Debit / Credit) in {sheet_name!r}"
        )

    header_cells = [str(c).strip().lower() if c else "" for c in rows[table_header_idx]]
    col = {name: header_cells.index(name) for name in header_cells if name}

    lines: list[dict] = []
    for row in rows[table_header_idx + 1 :]:
        if not row or not row[col.get("account", 0)]:
            continue
        acct = row[col["account"]]
        if str(acct).lower().startswith("total"):
            break
        ln = {
            "account": str(acct).strip(),
            "memo": str(row[col["memo"]]).strip() if "memo" in col and row[col["memo"]] else "",
        }
        if "subsidiary" in col and row[col["subsidiary"]]:
            ln["subsidiary"] = str(row[col["subsidiary"]]).strip()
        if "debit" in col and row[col["debit"]] not in (None, "", 0):
            ln["debit"] = float(row[col["debit"]])
        if "credit" in col and row[col["credit"]] not in (None, "", 0):
            ln["credit"] = float(row[col["credit"]])
        lines.append(ln)

    if not header_memo or not header_date:
        raise ValueError(
            "Could not extract Memo and Date from header rows. "
            "Provide them manually or label cells 'Memo:' and 'Date:'."
        )

    task_num = (header_memo.split(" - ")[0] if " - " in header_memo else header_memo).strip()

    return {
        "task_id": task_num,
        "memo": str(header_memo).strip(),
        "tran_date": str(header_date)[:10],
        "subsidiary": str(header_sub).strip() if header_sub else "",
        "lines": lines,
        "workpaper": path,
    }


def from_legacy_payload(legacy: dict) -> dict:
    """
    Convert a <your-automation>-style {recordType, data} payload back into the handoff
    format, so existing automations can pipe through this skill without rewrite.
    """
    if "data" not in legacy:
        raise ValueError("Expected a {recordType, data} legacy payload.")
    d = legacy["data"]

    lines = []
    for item in (d.get("line", {}) or {}).get("items", []):
        ln = {
            "account": (item.get("account") or {}).get("id"),
            "memo": item.get("memo"),
        }
        if item.get("debit") is not None:
            ln["debit"] = item["debit"]
        if item.get("credit") is not None:
            ln["credit"] = item["credit"]
        if (item.get("lineSubsidiary") or {}).get("id"):
            ln["subsidiary"] = item["lineSubsidiary"]["id"]
        for opt in ("department", "class", "location", "entity"):
            ref = item.get(opt)
            if ref and ref.get("id"):
                ln[opt] = ref["id"]
        lines.append(ln)

    memo = d.get("memo", "")
    task_num = memo.split(" - ")[0].strip() if " - " in memo else memo

    return {
        "task_id": task_num,
        "memo": memo,
        "tran_date": d.get("tranDate"),
        "subsidiary": (d.get("subsidiary") or {}).get("id"),
        "custom_form": (d.get("customForm") or {}).get("id"),
        "currency": (d.get("currency") or {}).get("id"),
        "lines": lines,
    }


# --- CLI ---
if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) != 2:
        print("usage: post_je.py <handoff.json | workbook.xlsx>", file=sys.stderr)
        sys.exit(2)

    src = sys.argv[1]
    if src.lower().endswith(".xlsx"):
        h = from_excel(src)
    else:
        with open(src, encoding="utf-8") as f:
            h = json.load(f)

    payload = to_ns_payload(h)
    print(json.dumps(payload, indent=2))
